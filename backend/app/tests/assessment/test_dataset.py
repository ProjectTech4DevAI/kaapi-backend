"""Tests for assessment/dataset.py upload and row counting behavior."""

from unittest.mock import MagicMock, patch

import pytest
from fastapi import HTTPException
from openpyxl.utils.exceptions import InvalidFileException

from app.services.assessment.dataset import (
    _count_csv_rows,
    _count_excel_rows,
    _count_rows,
    _preview_csv,
    _preview_excel,
    preview_dataset,
    upload_dataset,
)


class TestCountRows:
    def test_legacy_xls_rejected(self) -> None:
        with pytest.raises(ValueError, match="Legacy Excel format"):
            _count_rows(b"legacy-xls-content", ".xls")

    def test_count_excel_rows_invalid_file_re_raises(self) -> None:
        with patch(
            "openpyxl.load_workbook",
            side_effect=InvalidFileException("bad xlsx"),
        ):
            with pytest.raises(InvalidFileException):
                _count_excel_rows(b"bad")

    def test_count_excel_rows_unexpected_error_raises_value_error(self) -> None:
        with patch("openpyxl.load_workbook", side_effect=RuntimeError("boom")):
            with pytest.raises(ValueError, match="Failed to parse XLSX file"):
                _count_excel_rows(b"bad")

    def test_count_csv_rows(self) -> None:
        assert _count_csv_rows(b"a,b\n1,2\n\n3,4\n") == 2

    def test_count_rows_csv_and_xlsx(self) -> None:
        with patch("app.services.assessment.dataset._count_excel_rows", return_value=5):
            assert _count_rows(b"x", ".xlsx") == 5
        assert _count_rows(b"a,b\n1,2\n", ".csv") == 1


class TestUploadDataset:
    def test_invalid_xlsx_returns_422(self) -> None:
        session = MagicMock()
        with patch(
            "app.services.assessment.dataset.sanitize_dataset_name", return_value="ds-1"
        ), patch(
            "app.services.assessment.dataset._count_rows",
            side_effect=InvalidFileException("bad xlsx"),
        ):
            with pytest.raises(HTTPException) as exc_info:
                upload_dataset(
                    session=session,
                    file_content=b"invalid-xlsx",
                    file_ext=".xlsx",
                    dataset_name="ds-1",
                    description=None,
                    organization_id=1,
                    project_id=1,
                )
        assert exc_info.value.status_code == 422
        assert "Invalid XLSX file content" in exc_info.value.detail

    def test_count_rows_value_error_returns_422(self) -> None:
        session = MagicMock()
        with patch(
            "app.services.assessment.dataset.sanitize_dataset_name", return_value="ds-1"
        ), patch(
            "app.services.assessment.dataset._count_rows",
            side_effect=ValueError("Legacy Excel format (.xls) is not supported."),
        ):
            with pytest.raises(HTTPException) as exc_info:
                upload_dataset(
                    session=session,
                    file_content=b"bad",
                    file_ext=".xls",
                    dataset_name="ds-1",
                    description=None,
                    organization_id=1,
                    project_id=1,
                )
        assert exc_info.value.status_code == 422
        assert "Legacy Excel format" in exc_info.value.detail

    def test_count_rows_unexpected_error_returns_generic_422(self) -> None:
        session = MagicMock()
        with patch(
            "app.services.assessment.dataset.sanitize_dataset_name", return_value="ds-1"
        ), patch(
            "app.services.assessment.dataset._count_rows",
            side_effect=RuntimeError("unexpected"),
        ):
            with pytest.raises(HTTPException) as exc_info:
                upload_dataset(
                    session=session,
                    file_content=b"bad",
                    file_ext=".xlsx",
                    dataset_name="ds-1",
                    description=None,
                    organization_id=1,
                    project_id=1,
                )
        assert exc_info.value.status_code == 422
        assert "Unable to parse dataset file" in exc_info.value.detail

    def test_upload_dataset_success(self) -> None:
        session = MagicMock()
        created = MagicMock()
        created.id = 9
        with patch(
            "app.services.assessment.dataset.sanitize_dataset_name", return_value="ds-1"
        ), patch("app.services.assessment.dataset._count_rows", return_value=2), patch(
            "app.services.assessment.dataset._upload_file_to_object_store",
            return_value="s3://datasets/file.csv",
        ), patch(
            "app.services.assessment.dataset.create_assessment_dataset",
            return_value=created,
        ) as create_ds:
            result = upload_dataset(
                session=session,
                file_content=b"a,b\n1,2\n",
                file_ext=".csv",
                dataset_name="ds-1",
                description="desc",
                organization_id=1,
                project_id=1,
            )
        assert result.id == 9
        create_ds.assert_called_once()
        assert create_ds.call_args.kwargs["dataset_metadata"]["total_items_count"] == 2

    def test_preview_csv_returns_headers_and_rows(self) -> None:
        headers, rows = _preview_csv(b"a,b\n1,2\n\n3,4\n5,6\n", limit=2)
        assert headers == ["a", "b"]
        assert rows == [["1", "2"], ["3", "4"]]

    def test_preview_csv_handles_latin1_fallback(self) -> None:
        # \xff is invalid utf-8 -> falls back to latin-1
        headers, rows = _preview_csv(b"name\nca\xfffe\n", limit=5)
        assert headers == ["name"]
        assert rows and rows[0][0].startswith("ca")

    def test_preview_excel_returns_headers_and_rows(self) -> None:
        import io

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["x", "y"])
        ws.append([1, 2])
        ws.append([None, None])
        ws.append([3, 4])
        buf = io.BytesIO()
        wb.save(buf)
        headers, rows = _preview_excel(buf.getvalue(), limit=10)
        assert headers == ["x", "y"]
        assert rows == [["1", "2"], ["3", "4"]]

    def test_preview_excel_empty_workbook(self) -> None:
        import io

        import openpyxl

        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        headers, rows = _preview_excel(buf.getvalue(), limit=10)
        assert headers == [""] or headers == []
        assert rows == []

    def test_preview_dataset_missing_url_returns_404(self) -> None:
        ds = MagicMock()
        ds.object_store_url = None
        with pytest.raises(HTTPException) as exc_info:
            preview_dataset(session=MagicMock(), dataset=ds, project_id=1, limit=10)
        assert exc_info.value.status_code == 404

    def test_preview_dataset_missing_extension_returns_422(self) -> None:
        ds = MagicMock()
        ds.object_store_url = "s3://x"
        ds.dataset_metadata = {}
        with pytest.raises(HTTPException) as exc_info:
            preview_dataset(session=MagicMock(), dataset=ds, project_id=1, limit=10)
        assert exc_info.value.status_code == 422
        assert "Unsupported or missing" in exc_info.value.detail

    def test_preview_dataset_unknown_extension_returns_422(self) -> None:
        ds = MagicMock()
        ds.object_store_url = "s3://x"
        ds.dataset_metadata = {"file_extension": ".json"}
        with pytest.raises(HTTPException) as exc_info:
            preview_dataset(session=MagicMock(), dataset=ds, project_id=1, limit=10)
        assert exc_info.value.status_code == 422

    def test_preview_dataset_normalizes_extension_case(self) -> None:
        ds = MagicMock()
        ds.object_store_url = "s3://x"
        ds.dataset_metadata = {"file_extension": " .CSV "}
        storage = MagicMock()
        storage.get.return_value = b"a,b\n1,2\n"
        with patch(
            "app.services.assessment.dataset.get_cloud_storage", return_value=storage
        ):
            headers, rows = preview_dataset(
                session=MagicMock(), dataset=ds, project_id=1, limit=10
            )
        assert headers == ["a", "b"]
        assert rows == [["1", "2"]]

    def test_preview_dataset_legacy_xls_returns_422(self) -> None:
        ds = MagicMock()
        ds.object_store_url = "s3://x"
        ds.dataset_metadata = {"file_extension": ".xls"}
        with pytest.raises(HTTPException) as exc_info:
            preview_dataset(session=MagicMock(), dataset=ds, project_id=1, limit=10)
        assert exc_info.value.status_code == 422

    def test_preview_dataset_storage_failure_returns_502(self) -> None:
        ds = MagicMock()
        ds.object_store_url = "s3://x"
        ds.dataset_metadata = {"file_extension": ".csv"}
        storage = MagicMock()
        storage.get.side_effect = RuntimeError("boom")
        with patch(
            "app.services.assessment.dataset.get_cloud_storage", return_value=storage
        ):
            with pytest.raises(HTTPException) as exc_info:
                preview_dataset(session=MagicMock(), dataset=ds, project_id=1, limit=10)
        assert exc_info.value.status_code == 502

    def test_preview_dataset_invalid_xlsx_returns_422(self) -> None:
        ds = MagicMock()
        ds.object_store_url = "s3://x"
        ds.dataset_metadata = {"file_extension": ".xlsx"}
        storage = MagicMock()
        storage.get.return_value = b"not-a-real-xlsx"
        with patch(
            "app.services.assessment.dataset.get_cloud_storage", return_value=storage
        ), patch(
            "app.services.assessment.dataset._preview_excel",
            side_effect=InvalidFileException("bad"),
        ):
            with pytest.raises(HTTPException) as exc_info:
                preview_dataset(session=MagicMock(), dataset=ds, project_id=1, limit=10)
        assert exc_info.value.status_code == 422
        assert "Invalid XLSX" in exc_info.value.detail

    def test_preview_dataset_parse_error_returns_422(self) -> None:
        ds = MagicMock()
        ds.object_store_url = "s3://x"
        ds.dataset_metadata = {"file_extension": ".csv"}
        storage = MagicMock()
        storage.get.return_value = b"a,b\n1,2\n"
        with patch(
            "app.services.assessment.dataset.get_cloud_storage", return_value=storage
        ), patch(
            "app.services.assessment.dataset._preview_csv",
            side_effect=RuntimeError("boom"),
        ):
            with pytest.raises(HTTPException) as exc_info:
                preview_dataset(session=MagicMock(), dataset=ds, project_id=1, limit=10)
        assert exc_info.value.status_code == 422
        assert "Unable to parse" in exc_info.value.detail

    def test_preview_dataset_csv_success(self) -> None:
        ds = MagicMock()
        ds.object_store_url = "s3://x"
        ds.dataset_metadata = {"file_extension": ".csv"}
        storage = MagicMock()
        storage.get.return_value = b"a,b\n1,2\n3,4\n"
        with patch(
            "app.services.assessment.dataset.get_cloud_storage", return_value=storage
        ):
            headers, rows = preview_dataset(
                session=MagicMock(), dataset=ds, project_id=1, limit=10
            )
        assert headers == ["a", "b"]
        assert rows == [["1", "2"], ["3", "4"]]

    def test_upload_dataset_object_store_failure_returns_500(self) -> None:
        session = MagicMock()
        with patch(
            "app.services.assessment.dataset.sanitize_dataset_name", return_value="ds-1"
        ), patch("app.services.assessment.dataset._count_rows", return_value=1), patch(
            "app.services.assessment.dataset._upload_file_to_object_store",
            return_value=None,
        ):
            with pytest.raises(HTTPException) as exc_info:
                upload_dataset(
                    session=session,
                    file_content=b"a,b\n1,2\n",
                    file_ext=".csv",
                    dataset_name="ds-1",
                    description=None,
                    organization_id=1,
                    project_id=1,
                )
        assert exc_info.value.status_code == 500
