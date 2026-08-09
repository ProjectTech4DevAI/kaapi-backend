"""Tests for assessment/batch.py provider routing in submit_assessment_batch."""

import io
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.crud.assessment.batch import (
    _build_text_prompt,
    _load_dataset_rows,
    _parse_excel_rows,
    build_anthropic_jsonl,
    build_google_jsonl,
    build_openai_jsonl,
    submit_assessment_batch,
)
from app.models.assessment import AssessmentAttachment
from app.models.llm.constants import DEFAULT_ASSESSMENT_BATCH_MAX_TOKENS
from app.services.assessment.utils.attachments import (
    _guess_image_mime_from_url,
    attachment_type_for_row,
    build_anthropic_attachment_parts,
    build_gemini_attachment_parts,
    resolve_attachment_values,
    resolve_item_type,
    split_attachment_urls,
    to_direct_attachment_url,
)


def _make_run() -> MagicMock:
    run = MagicMock()
    run.id = 99
    return run


def _make_assessment() -> MagicMock:
    assessment = MagicMock()
    assessment.id = 21
    assessment.experiment_name = "exp-v1"
    return assessment


def _make_dataset() -> MagicMock:
    dataset = MagicMock()
    dataset.id = 8
    return dataset


class TestSubmitAssessmentBatchProviderRouting:
    def test_openai_native_routes_to_openai_batch(self) -> None:
        session = MagicMock()
        run = _make_run()
        dataset = _make_dataset()
        config_blob = SimpleNamespace(
            completion=SimpleNamespace(
                provider="openai-native",
                params={"instructions": "config system"},
            )
        )
        batch_job = MagicMock()
        batch_job.id = 1
        batch_job.total_items = 1

        with (
            patch(
                "app.crud.assessment.batch._load_dataset_rows",
                return_value=[{"question": "q1"}],
            ),
            patch(
                "app.crud.assessment.batch.map_kaapi_to_openai_params",
                return_value=({}, []),
            ) as map_params,
            patch(
                "app.crud.assessment.batch.build_openai_jsonl",
                return_value=[{"custom_id": "row_0"}],
            ),
            patch(
                "app.crud.assessment.batch.get_openai_client",
                return_value=MagicMock(),
            ),
            patch(
                "app.crud.assessment.batch.OpenAIBatchProvider",
                return_value=MagicMock(),
            ),
            patch(
                "app.crud.assessment.batch.start_batch_job",
                return_value=batch_job,
            ) as start_batch,
        ):
            result = submit_assessment_batch(
                session=session,
                run=run,
                assessment=_make_assessment(),
                dataset=dataset,
                config_blob=config_blob,
                assessment_input={
                    "text_columns": ["question"],
                    "attachments": [],
                    "system_instruction": "request system",
                },
                organization_id=1,
                project_id=1,
            )

        assert result.id == 1
        assert map_params.call_args.kwargs["session"] is session
        # Instruction now comes from the resolved config blob; the request has no
        # system_instruction override anymore.
        assert map_params.call_args.kwargs["kaapi_params"]["instructions"] == (
            "config system"
        )
        assert start_batch.call_args.kwargs["provider_name"] == "openai"

    def test_config_instruction_is_used(self) -> None:
        session = MagicMock()
        run = _make_run()
        dataset = _make_dataset()
        config_blob = SimpleNamespace(
            completion=SimpleNamespace(
                provider="openai",
                params={"instructions": "config system", "model": "gpt-4.1-mini"},
            )
        )
        batch_job = MagicMock()
        batch_job.id = 3
        batch_job.total_items = 1

        with (
            patch(
                "app.crud.assessment.batch._load_dataset_rows",
                return_value=[{"question": "q1"}],
            ),
            patch(
                "app.crud.assessment.batch.map_kaapi_to_openai_params",
                return_value=({"model": "gpt-4.1-mini"}, []),
            ) as map_params,
            patch(
                "app.crud.assessment.batch.build_openai_jsonl",
                return_value=[{"custom_id": "row_0"}],
            ),
            patch(
                "app.crud.assessment.batch.get_openai_client",
                return_value=MagicMock(),
            ),
            patch(
                "app.crud.assessment.batch.OpenAIBatchProvider",
                return_value=MagicMock(),
            ),
            patch(
                "app.crud.assessment.batch.start_batch_job",
                return_value=batch_job,
            ),
        ):
            submit_assessment_batch(
                session=session,
                run=run,
                assessment=_make_assessment(),
                dataset=dataset,
                config_blob=config_blob,
                assessment_input={"text_columns": ["question"], "attachments": []},
                organization_id=1,
                project_id=1,
            )

        assert map_params.call_args.kwargs["session"] is session
        assert (
            map_params.call_args.kwargs["kaapi_params"]["instructions"]
            == "config system"
        )

    def test_google_native_routes_to_google_batch(self) -> None:
        session = MagicMock()
        run = _make_run()
        dataset = _make_dataset()
        config_blob = SimpleNamespace(
            completion=SimpleNamespace(
                provider="google-native",
                params={"instructions": "config system"},
            )
        )
        batch_job = MagicMock()
        batch_job.id = 2
        batch_job.total_items = 1
        gemini_client = MagicMock()
        gemini_client.client = MagicMock()

        with (
            patch(
                "app.crud.assessment.batch._load_dataset_rows",
                return_value=[{"question": "q1"}],
            ),
            patch(
                "app.crud.assessment.batch.map_kaapi_to_google_params",
                return_value=({"model": "gemini-2.5-pro"}, []),
            ) as map_params,
            patch(
                "app.crud.assessment.batch.build_google_jsonl",
                return_value=[{"key": "row_0"}],
            ),
            patch("app.crud.assessment.batch.GeminiClient") as gemini_cls,
            patch(
                "app.crud.assessment.batch.GeminiBatchProvider",
                return_value=MagicMock(),
            ),
            patch(
                "app.crud.assessment.batch.start_batch_job",
                return_value=batch_job,
            ) as start_batch,
        ):
            gemini_cls.from_credentials.return_value = gemini_client
            result = submit_assessment_batch(
                session=session,
                run=run,
                assessment=_make_assessment(),
                dataset=dataset,
                config_blob=config_blob,
                assessment_input={
                    "text_columns": ["question"],
                    "attachments": [],
                    "system_instruction": "request system",
                },
                organization_id=1,
                project_id=1,
            )

        assert result.id == 2
        assert map_params.call_args.args[0]["instructions"] == "config system"
        assert start_batch.call_args.kwargs["provider_name"] == "google"

    def test_anthropic_native_routes_to_anthropic_batch(self) -> None:
        session = MagicMock()
        run = _make_run()
        dataset = _make_dataset()
        config_blob = SimpleNamespace(
            completion=SimpleNamespace(
                provider="anthropic-native",
                params={"instructions": "config system"},
            )
        )
        batch_job = MagicMock()
        batch_job.id = 4
        batch_job.total_items = 1

        with (
            patch(
                "app.crud.assessment.batch._load_dataset_rows",
                return_value=[{"question": "q1"}],
            ),
            patch(
                "app.crud.assessment.batch.map_kaapi_to_anthropic_params",
                return_value=({"model": "claude-sonnet-4-6"}, []),
            ) as map_params,
            patch(
                "app.crud.assessment.batch.build_anthropic_jsonl",
                return_value=[{"custom_id": "row_0"}],
            ),
            patch(
                "app.crud.assessment.batch.get_anthropic_client",
                return_value=MagicMock(),
            ),
            patch(
                "app.crud.assessment.batch.AnthropicBatchProvider",
                return_value=MagicMock(),
            ),
            patch(
                "app.crud.assessment.batch.start_batch_job",
                return_value=batch_job,
            ) as start_batch,
        ):
            result = submit_assessment_batch(
                session=session,
                run=run,
                assessment=_make_assessment(),
                dataset=dataset,
                config_blob=config_blob,
                assessment_input={
                    "text_columns": ["question"],
                    "attachments": [],
                    "system_instruction": "request system",
                },
                organization_id=1,
                project_id=1,
            )

        assert result.id == 4
        assert map_params.call_args.args[0]["instructions"] == "config system"
        assert start_batch.call_args.kwargs["provider_name"] == "anthropic"
        assert start_batch.call_args.kwargs["config"]["model"] == "claude-sonnet-4-6"
        assert start_batch.call_args.kwargs["config"]["max_tokens"] == (
            DEFAULT_ASSESSMENT_BATCH_MAX_TOKENS
        )


class TestBatchDatasetParsing:
    def test_load_dataset_rows_routes_xlsx_to_excel_parser(self) -> None:
        session = MagicMock()
        dataset = MagicMock()
        dataset.id = 8
        dataset.project_id = 1
        dataset.object_store_url = "s3://bucket/key"
        dataset.dataset_metadata = {"file_extension": ".xlsx"}

        storage = MagicMock()
        stream_body = MagicMock()
        stream_body.read.return_value = b"xlsx-content"
        storage.stream.return_value = stream_body

        expected = [{"question": "q1"}]
        with (
            patch("app.crud.assessment.batch.get_cloud_storage", return_value=storage),
            patch(
                "app.crud.assessment.batch._parse_excel_rows",
                return_value=expected,
            ) as parse_excel,
        ):
            result = _load_dataset_rows(session=session, dataset=dataset)

        assert result == expected
        parse_excel.assert_called_once_with(b"xlsx-content")

    def test_load_dataset_rows_rejects_legacy_xls(self) -> None:
        session = MagicMock()
        dataset = MagicMock()
        dataset.id = 8
        dataset.project_id = 1
        dataset.object_store_url = "s3://bucket/key"
        dataset.dataset_metadata = {"file_extension": ".xls"}

        storage = MagicMock()
        stream_body = MagicMock()
        stream_body.read.return_value = b"legacy-xls-content"
        storage.stream.return_value = stream_body

        with patch("app.crud.assessment.batch.get_cloud_storage", return_value=storage):
            with pytest.raises(ValueError, match="Legacy Excel format"):
                _load_dataset_rows(session=session, dataset=dataset)

    def test_parse_excel_rows_invalid_payload_raises(self) -> None:
        with pytest.raises((ValueError, InvalidFileException)):
            _parse_excel_rows(b"not-a-valid-xlsx")

    def test_parse_excel_rows_success(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["question", "answer"])
        ws.append(["What is 2+2?", "4"])
        ws.append(["", None])  # empty row should be skipped
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()

        rows = _parse_excel_rows(buf.getvalue())
        assert rows == [{"question": "What is 2+2?", "answer": "4"}]

    def test_parse_excel_rows_returns_empty_when_sheet_missing(self) -> None:
        fake_wb = MagicMock()
        fake_wb.active = None
        with patch(
            "app.crud.assessment.batch.openpyxl.load_workbook", return_value=fake_wb
        ):
            assert _parse_excel_rows(b"irrelevant") == []
        fake_wb.close.assert_called_once()

    def test_parse_excel_rows_returns_empty_when_header_missing(self) -> None:
        fake_ws = MagicMock()
        fake_ws.iter_rows.return_value = iter([])
        fake_wb = MagicMock()
        fake_wb.active = fake_ws
        with patch(
            "app.crud.assessment.batch.openpyxl.load_workbook", return_value=fake_wb
        ):
            assert _parse_excel_rows(b"irrelevant") == []
        fake_wb.close.assert_called_once()

    def test_parse_excel_rows_invalid_file_exception_re_raises(self) -> None:
        with patch(
            "app.crud.assessment.batch.openpyxl.load_workbook",
            side_effect=InvalidFileException("bad xlsx"),
        ):
            with pytest.raises(InvalidFileException):
                _parse_excel_rows(b"bad")

    def test_parse_excel_rows_unexpected_exception_raises_value_error(self) -> None:
        with patch(
            "app.crud.assessment.batch.openpyxl.load_workbook",
            side_effect=RuntimeError("boom"),
        ):
            with pytest.raises(ValueError, match="Failed to parse XLSX dataset rows"):
                _parse_excel_rows(b"bad")


class TestBatchHelpers:
    def test_build_text_prompt_template_and_concat(self) -> None:
        row = {"q": " What? ", "ctx": "Context"}
        templated = _build_text_prompt(row, ["q", "ctx"], "Q:{q}\nC:{ctx}")
        assert "Q:" in templated
        assert "What?" in templated
        concatenated = _build_text_prompt(row, ["q", "ctx"], None)
        assert "What?" in concatenated
        assert concatenated.endswith("\nContext")

    def test_split_and_direct_urls(self) -> None:
        urls = split_attachment_urls(" https://a.com\nhttps://b.com , https://c.com ")
        assert urls == ["https://a.com", "https://b.com", "https://c.com"]
        image_url = to_direct_attachment_url(
            "https://drive.google.com/file/d/abc123/view?usp=sharing", "image"
        )
        assert "googleusercontent.com" in image_url
        pdf_url = to_direct_attachment_url(
            "https://drive.google.com/open?id=abc123", "pdf"
        )
        assert "drive.google.com/uc" in pdf_url

    def test_url_mime_guessers(self) -> None:
        assert _guess_image_mime_from_url("https://x/y/file.jpeg") == "image/jpeg"
        assert _guess_image_mime_from_url("https://x/y/file.unknown") is None

    def testresolve_attachment_values(self) -> None:
        image_url_att = AssessmentAttachment(column="img", type="image", format="url")
        pdf_url_att = AssessmentAttachment(column="pdf", type="pdf", format="url")

        values = resolve_attachment_values(
            "https://example.com/a.png,https://example.com/b.png", image_url_att
        )
        assert len(values) == 2
        assert values[0]["type"] == "input_image"
        assert values[0]["image_url"] == "https://example.com/a.png"

        values = resolve_attachment_values("https://example.com/a.pdf", pdf_url_att)
        assert values[0]["type"] == "input_file"
        assert values[0]["file_url"] == "https://example.com/a.pdf"

    def test_build_openai_and_google_jsonl(self) -> None:
        rows = [{"q": "What is 2+2?", "img": "https://example.com/a.png"}]
        attachments = [AssessmentAttachment(column="img", type="image", format="url")]

        openai_jsonl = build_openai_jsonl(
            rows=rows,
            text_columns=["q"],
            attachments=attachments,
            prompt_template=None,
            openai_params={"model": "gpt-4.1-mini"},
        )
        assert len(openai_jsonl) == 1
        assert openai_jsonl[0]["custom_id"] == "row_0"

        google_jsonl = build_google_jsonl(
            rows=rows,
            text_columns=["q"],
            attachments=attachments,
            prompt_template=None,
            google_params={"temperature": 0.2, "instructions": "system"},
        )
        assert len(google_jsonl) == 1
        assert google_jsonl[0]["key"] == "row_0"
        assert google_jsonl[0]["request"]["systemInstruction"] == {
            "parts": [{"text": "system"}]
        }

    def test_build_anthropic_jsonl(self) -> None:
        rows = [
            {
                "q": "What is 2+2?",
                "img": "https://example.com/a.png",
                "pdf": "https://example.com/d.pdf",
            }
        ]
        attachments = [
            AssessmentAttachment(column="img", type="image", format="url"),
            AssessmentAttachment(column="pdf", type="pdf", format="url"),
        ]

        jsonl = build_anthropic_jsonl(
            rows=rows,
            text_columns=["q"],
            attachments=attachments,
            prompt_template=None,
            anthropic_params={"model": "claude-sonnet-4-6", "system": "be brief"},
        )
        assert len(jsonl) == 1
        assert jsonl[0]["custom_id"] == "row_0"
        params = jsonl[0]["params"]
        assert params["model"] == "claude-sonnet-4-6"
        assert params["system"] == "be brief"
        content = params["messages"][0]["content"]
        assert content[0] == {"type": "text", "text": "What is 2+2?"}
        assert content[1] == {
            "type": "image",
            "source": {"type": "url", "url": "https://example.com/a.png"},
        }
        assert content[2] == {
            "type": "document",
            "source": {"type": "url", "url": "https://example.com/d.pdf"},
        }

    def test_build_anthropic_jsonl_skips_empty_rows(self) -> None:
        jsonl = build_anthropic_jsonl(
            rows=[{"q": "  "}],
            text_columns=["q"],
            attachments=[],
            prompt_template=None,
            anthropic_params={"model": "claude-sonnet-4-6"},
        )
        assert jsonl == []

    def test_build_anthropic_jsonl_uses_row_indices(self) -> None:
        jsonl = build_anthropic_jsonl(
            rows=[{"q": "a"}, {"q": "b"}],
            text_columns=["q"],
            attachments=[],
            prompt_template=None,
            anthropic_params={"model": "claude-sonnet-4-6"},
            row_indices=[3, 7],
        )
        assert [line["custom_id"] for line in jsonl] == ["row_3", "row_7"]

    def test_build_anthropic_attachment_parts(self) -> None:
        image_att = AssessmentAttachment(column="img", type="image", format="url")
        parts = build_anthropic_attachment_parts(
            "https://example.com/a.png,https://example.com/b.png", image_att
        )
        assert len(parts) == 2
        assert parts[0]["type"] == "image"
        assert parts[0]["source"] == {"type": "url", "url": "https://example.com/a.png"}


class TestResolveItemType:
    """Image/pdf routing now trusts the user-declared type (no detection)."""

    def test_declared_image(self) -> None:
        assert resolve_item_type("image") == "image"

    def test_declared_pdf(self) -> None:
        assert resolve_item_type("pdf") == "pdf"

    def test_override_wins(self) -> None:
        assert resolve_item_type("image", "pdf") == "pdf"
        assert resolve_item_type("pdf", "image") == "image"

    def test_mixed_without_override_is_unresolved(self) -> None:
        assert resolve_item_type("mixed") is None

    def test_unknown_declared_is_unresolved(self) -> None:
        assert resolve_item_type("whatever") is None

    def test_column_uses_single_declared_type(self) -> None:
        """One column, many URLs -> all routed by the declared type."""
        att = AssessmentAttachment(column="docs", type="pdf", format="url")
        value = "https://x.com/a/photo.jpg, https://x.com/b/report"
        resolved = resolve_attachment_values(value, att)
        types = [obj["type"] for obj in resolved]
        assert types == ["input_file", "input_file"]


class TestAttachmentMime:
    def test_guess_image_mime_from_url_variants(self) -> None:
        assert _guess_image_mime_from_url("http://x/a.PNG") == "image/png"
        assert _guess_image_mime_from_url("http://x/a.jpeg") == "image/jpeg"
        assert _guess_image_mime_from_url("http://x/a.webp") == "image/webp"
        assert _guess_image_mime_from_url("http://x/a.txt") is None


class TestAttachmentTypeForRow:
    def test_mixed_resolves_from_type_column(self) -> None:
        from app.services.assessment.utils.attachments import attachment_type_for_row

        att = AssessmentAttachment(
            column="Docs",
            type="mixed",
            format="url",
            type_column="DOC type",
            type_value_map={"Photo": "image", "Report": "pdf"},
        )
        assert attachment_type_for_row(att, {"DOC type": "Photo"}) == "image"
        assert attachment_type_for_row(att, {"DOC type": "Report"}) == "pdf"
        assert attachment_type_for_row(att, {"DOC type": "Unknown"}) is None

    def test_mixed_resolves_comma_separated_value_lists(self) -> None:
        from app.services.assessment.utils.attachments import attachment_type_for_row

        att = AssessmentAttachment(
            column="Docs",
            type="mixed",
            format="url",
            type_column="DOC type",
            type_value_map={"Img-Prototype, Img-Handtext": "image", "Pdf": "pdf"},
        )

        assert attachment_type_for_row(att, {"DOC type": "Img-Prototype"}) == "image"
        assert attachment_type_for_row(att, {"DOC type": "Img-Handtext"}) == "image"
        assert attachment_type_for_row(att, {"DOC type": "pdf"}) == "pdf"

    def test_mixed_resolves_row_value_lists_when_same_type(self) -> None:
        from app.services.assessment.utils.attachments import attachment_type_for_row

        att = AssessmentAttachment(
            column="Docs",
            type="mixed",
            format="url",
            type_column="DOC type",
            type_value_map={"Img-Prototype, Img-Handtext": "image", "Pdf": "pdf"},
        )

        assert (
            attachment_type_for_row(
                att,
                {"DOC type": "Img-Prototype, Img-Handtext"},
            )
            == "image"
        )
        assert attachment_type_for_row(att, {"DOC type": "Img-Prototype, Pdf"}) is None

    def test_mixed_missing_type_mapping_fields_returns_none(self) -> None:
        from app.services.assessment.utils.attachments import attachment_type_for_row

        att = SimpleNamespace(column="Docs", type="mixed", format="url")

        assert attachment_type_for_row(att, {"Docs": "x"}) is None

    def test_non_mixed_returns_none(self) -> None:
        from app.services.assessment.utils.attachments import attachment_type_for_row

        att = AssessmentAttachment(column="Docs", type="image", format="url")
        assert attachment_type_for_row(att, {"Docs": "x"}) is None

    def test_mixed_config_missing_routing_fields_is_rejected(self) -> None:
        import pytest
        from pydantic import ValidationError

        with pytest.raises(ValidationError):
            AssessmentAttachment(column="Docs", type="mixed", format="url")

    def test_mixed_config_invalid_map_value_is_rejected(self) -> None:
        import pytest
        from pydantic import ValidationError

        with pytest.raises(ValidationError):
            AssessmentAttachment(
                column="Docs",
                type="mixed",
                format="url",
                type_column="DOC type",
                type_value_map={"Report": "spreadsheet"},
            )

    def test_override_forces_part_type(self) -> None:
        from app.services.assessment.utils.attachments import resolve_attachment_values

        att = AssessmentAttachment(
            column="Docs",
            type="mixed",
            format="url",
            type_column="DOC type",
            type_value_map={"Report": "pdf"},
        )
        url = "https://drive.google.com/file/d/ID/view"
        parts = resolve_attachment_values(url, att, type_override="pdf")
        assert parts[0]["type"] == "input_file"


class TestAttachmentResolutionBranches:
    _IMG = AssessmentAttachment(column="Docs", type="image", format="url")
    _PDF = AssessmentAttachment(column="Docs", type="pdf", format="url")
    _MIXED = AssessmentAttachment(
        column="Docs",
        type="mixed",
        format="url",
        type_column="DOC type",
        type_value_map={"Report": "pdf"},
    )

    def test_blank_value_returns_empty(self) -> None:
        assert resolve_attachment_values("  ", self._IMG) == []
        assert build_gemini_attachment_parts("  ", self._IMG) == []

    def test_unresolved_mixed_is_skipped(self) -> None:
        url = "https://x.com/a.jpg"
        # No override and declared 'mixed' -> unresolved -> skip rather than guess.
        assert resolve_attachment_values(url, self._MIXED) == []
        assert build_gemini_attachment_parts(url, self._MIXED) == []

    def test_gemini_image_and_pdf_parts(self) -> None:
        img = build_gemini_attachment_parts("https://x.com/a.png", self._IMG)[0]
        pdf = build_gemini_attachment_parts("https://x.com/a.pdf", self._PDF)[0]
        assert img["fileData"]["mimeType"] == "image/png"
        assert pdf["fileData"]["mimeType"] == "application/pdf"

    def test_type_for_row_blank_value_returns_none(self) -> None:
        assert attachment_type_for_row(self._MIXED, {"DOC type": "  "}) is None

    def test_type_for_row_ignores_invalid_map_value(self) -> None:
        # SimpleNamespace bypasses the model validator to exercise the guard that
        # skips map entries whose target type isn't 'image'/'pdf'.
        att = SimpleNamespace(
            type="mixed",
            type_column="DOC type",
            type_value_map={"Report": "spreadsheet"},
        )
        assert attachment_type_for_row(att, {"DOC type": "Report"}) is None
