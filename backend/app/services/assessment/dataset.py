"""Dataset management service for assessments (CSV + XLSX).

Upload stores files directly to object store as-is (no column validation,
no format conversion). Row count is computed for metadata.
"""

import csv
import io
import logging

from fastapi import HTTPException
from sqlmodel import Session

from app.core.cloud import get_cloud_storage
from app.core.storage_utils import generate_timestamped_filename, upload_to_object_store
from app.crud.assessment.dataset import create_assessment_dataset
from app.models.evaluation import EvaluationDataset
from app.services.evaluations.validators import sanitize_dataset_name

logger = logging.getLogger(__name__)

try:
    from openpyxl.utils.exceptions import InvalidFileException
except Exception:  # pragma: no cover - openpyxl is expected in runtime deps

    class InvalidFileException(Exception):
        pass


_MIME_TYPES = {
    ".csv": "text/csv",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}


def _upload_file_to_object_store(
    session: Session,
    project_id: int,
    file_content: bytes,
    file_ext: str,
    dataset_name: str,
) -> str | None:
    """Upload the raw file to object store, preserving original format."""
    extension = file_ext.lstrip(".")
    filename = generate_timestamped_filename(dataset_name, extension=extension)
    content_type = _MIME_TYPES.get(file_ext, "application/octet-stream")

    try:
        storage = get_cloud_storage(session=session, project_id=project_id)
        return upload_to_object_store(
            storage=storage,
            content=file_content,
            filename=filename,
            subdirectory="datasets",
            content_type=content_type,
        )
    except Exception as e:
        logger.warning(
            f"[_upload_file_to_object_store] Failed to upload | {e}",
            exc_info=True,
        )
        return None


def _count_csv_rows(content: bytes) -> int:
    """Count data rows in a CSV file (excluding header)."""
    try:
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = content.decode(encoding)
                break
            except (UnicodeDecodeError, ValueError):
                continue
        else:
            text = content.decode("utf-8", errors="replace")

        reader = csv.reader(io.StringIO(text))
        next(reader, None)
        return sum(1 for row in reader if any(cell.strip() for cell in row))
    except Exception as e:
        logger.warning(f"[_count_csv_rows] Failed to count rows | {e}")
        return 0


def _count_excel_rows(content: bytes) -> int:
    """Count data rows in an Excel file (excluding header)."""
    wb = None
    try:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return 0

        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            return 0

        return sum(
            1 for row in rows_iter if row and any(cell is not None for cell in row)
        )
    except InvalidFileException as e:
        logger.warning("[_count_excel_rows] Invalid XLSX file content: %s", e)
        raise
    except Exception as e:
        logger.warning(
            "[_count_excel_rows] Failed to count rows | %s", e, exc_info=True
        )
        raise ValueError("Failed to parse XLSX file") from e
    finally:
        if wb is not None:
            wb.close()


def _count_rows(content: bytes, file_ext: str) -> int:
    """Count data rows in a file (CSV or XLSX), excluding the header."""
    if file_ext == ".xls":
        raise ValueError(
            "Legacy Excel format (.xls) is not supported. Please upload .xlsx or .csv."
        )
    if file_ext == ".xlsx":
        return _count_excel_rows(content)
    return _count_csv_rows(content)


def _stringify(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _preview_csv(content: bytes, limit: int) -> tuple[list[str], list[list[str]]]:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except (UnicodeDecodeError, ValueError):
            continue
    else:
        text = content.decode("utf-8", errors="replace")

    reader = csv.reader(io.StringIO(text))
    header = next(reader, None) or []
    headers = [_stringify(cell) for cell in header]

    rows: list[list[str]] = []
    for row in reader:
        if not any(cell.strip() for cell in row):
            continue
        rows.append([_stringify(cell) for cell in row])
        if len(rows) >= limit:
            break
    return headers, rows


def _preview_excel(content: bytes, limit: int) -> tuple[list[str], list[list[str]]]:
    import openpyxl

    wb = None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return [], []

        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None) or ()
        headers = [_stringify(cell) for cell in header]

        rows: list[list[str]] = []
        for row in rows_iter:
            if not row or not any(cell is not None for cell in row):
                continue
            rows.append([_stringify(cell) for cell in row])
            if len(rows) >= limit:
                break
        return headers, rows
    finally:
        if wb is not None:
            wb.close()


def preview_dataset(
    session: Session,
    dataset: EvaluationDataset,
    project_id: int,
    limit: int,
) -> tuple[list[str], list[list[str]]]:
    """Return the first `limit` data rows (plus header) of a dataset file."""
    if not dataset.object_store_url:
        raise HTTPException(
            status_code=404, detail="Dataset has no underlying file to preview."
        )

    raw_ext = (dataset.dataset_metadata or {}).get("file_extension")
    file_ext = raw_ext.strip().lower() if isinstance(raw_ext, str) else None
    if file_ext == ".xls":
        raise HTTPException(
            status_code=422,
            detail="Legacy Excel format (.xls) is not supported.",
        )
    if file_ext not in {".csv", ".xlsx"}:
        raise HTTPException(
            status_code=422,
            detail="Unsupported or missing file extension.",
        )

    storage = get_cloud_storage(session=session, project_id=project_id)
    try:
        content = storage.get(dataset.object_store_url)
    except Exception as e:
        logger.warning(
            f"[preview_dataset] Failed to fetch file | dataset_id={dataset.id} | {e}",
            exc_info=True,
        )
        raise HTTPException(
            status_code=502, detail="Failed to fetch dataset file from storage."
        ) from e

    try:
        if file_ext == ".xlsx":
            return _preview_excel(content, limit)
        return _preview_csv(content, limit)
    except InvalidFileException as e:
        raise HTTPException(status_code=422, detail="Invalid XLSX file content.") from e
    except Exception as e:
        logger.warning(
            f"[preview_dataset] Failed to parse file | dataset_id={dataset.id} | {e}",
            exc_info=True,
        )
        raise HTTPException(
            status_code=422, detail="Unable to parse dataset file for preview."
        ) from e


def upload_dataset(
    session: Session,
    file_content: bytes,
    file_ext: str,
    dataset_name: str,
    description: str | None,
    organization_id: int,
    project_id: int,
) -> EvaluationDataset:
    """Upload a dataset file directly to object store and record metadata."""
    original_name = dataset_name
    try:
        dataset_name = sanitize_dataset_name(dataset_name)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=f"Invalid dataset name: {str(e)}")

    if original_name != dataset_name:
        logger.info(
            f"[upload_dataset] Dataset name sanitized | '{original_name}' -> '{dataset_name}'"
        )

    try:
        row_count = _count_rows(file_content, file_ext)
    except InvalidFileException as e:
        raise HTTPException(
            status_code=422,
            detail="Invalid XLSX file content. Please upload a valid .xlsx file.",
        ) from e
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e)) from e
    except Exception as e:
        raise HTTPException(
            status_code=422,
            detail="Unable to parse dataset file. Please upload a valid CSV or XLSX file.",
        ) from e

    logger.info(
        f"[upload_dataset] Uploading dataset | dataset={dataset_name} | "
        f"file_type={file_ext} | rows={row_count} | "
        f"org_id={organization_id} | project_id={project_id}"
    )

    object_store_url = _upload_file_to_object_store(
        session=session,
        project_id=project_id,
        file_content=file_content,
        file_ext=file_ext,
        dataset_name=dataset_name,
    )
    if not object_store_url:
        logger.error(
            f"[upload_dataset] Object store upload failed | dataset={dataset_name} | "
            f"org_id={organization_id} | project_id={project_id}"
        )
        raise HTTPException(
            status_code=500,
            detail="Failed to upload dataset file. Please try again.",
        )

    metadata = {
        "file_extension": file_ext,
        "file_size_bytes": len(file_content),
        "total_items_count": row_count,
    }

    dataset = create_assessment_dataset(
        session=session,
        name=dataset_name,
        description=description,
        dataset_metadata=metadata,
        object_store_url=object_store_url,
        langfuse_dataset_id=None,
        organization_id=organization_id,
        project_id=project_id,
    )

    logger.info(
        f"[upload_dataset] Created dataset record | "
        f"id={dataset.id} | name={dataset_name} | rows={row_count}"
    )

    return dataset
